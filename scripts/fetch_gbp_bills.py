"""
fetch_gbp_bills.py
==================
Scraper for UK gilt nominal spot yields from Bank of England.

Source:   Bank of England — UK Government Liability Curve, nominal, daily
URL:      https://www.bankofengland.co.uk/-/media/boe/files/statistics/
              yield-curves/glcnominalddata.zip  (~38 MB)
Format:   ZIP archive containing multiple XLSX files split by year ranges:
              GLC_Nominal_daily_data_1979_to_1984.xlsx
              GLC_Nominal_daily_data_1985_to_1989.xlsx
              GLC_Nominal_daily_data_1990_to_1994.xlsx
              GLC_Nominal_daily_data_1995_to_1999.xlsx
              GLC_Nominal_daily_data_2000_to_2004.xlsx
              GLC_Nominal_daily_data_2005_to_2015.xlsx
              GLC_Nominal_daily_data_2016_to_2024.xlsx
              GLC_Nominal_daily_data_2025_to_present.xlsx

Sheets in each XLSX:
    info                — Bank of England documentation/disclaimer
    1. fwds, short end  — forward rates short end (months 1-60)
    2. fwd curve        — forward rates full curve (0.5Y-25Y)
    3. spot, short end  — spot rates short end (months 1-60)
    4. spot curve       — SPOT RATES FULL CURVE (0.5Y-25Y)  ← we use this

We use sheet "4. spot curve" because it contains the full curve with stable
yearly tenors (0.5Y, 1Y, 1.5Y, 2Y, ..., 25Y). The "short end" sheet has month
granularity but the first 5 months are empty (the gilt curve only stabilises
from 6M onwards).

Structure of sheet "4. spot curve":
    Row 1: title
    Row 2-3: header labels
    Row 4: "years:", 0.5, 1, 1.5, 2, 2.5, 3, ..., 25  ← maturity in years
    Row 5: error placeholder ("#VALUE!" or "Refresh") — skip
    Row 6+: date in col 0, yields in subsequent columns

Why 6M as bill_short_GBP:
    The BoE gilt nominal curve has 0.5-year (=6M) as its shortest stable point
    (the curve is fitted from gilts plus GC repo down to 1 week, but practical
    values stabilise from 6M up). This matches DGS6MO in fetch_us_bills.py,
    giving GBP a symmetric 6M/6M pair vs US (no curve-slope bias).

Why download the entire ZIP daily (rather than incremental updates):
    Simplicity and robustness. The full archive download is ~38 MB which takes
    ~10-15 seconds in GitHub Actions; trivial. Downloading the full archive
    each run also handles BoE backward revisions automatically (gilt curves
    are model-fitted and the BoE occasionally re-estimates historical data).

Performance note:
    Among the 8 XLSX files in the ZIP, we only need the files that cover the
    last 5 years (currently 2016-2024 + 2025-present). The 6 older files are
    skipped to save memory and time.

Output:
    data/GBP_BILL_6M.csv   — 6-month gilt nominal (bill_short_GBP for engine)
    data/GBP_BILL_1Y.csv   — 1-year
    data/GBP_BILL_2Y.csv   — 2-year
    data/GBP_BILL_5Y.csv   — 5-year   (curve cross-validation)
    data/GBP_BILL_10Y.csv  — 10-year  (benchmark)

License: Bank of England public statistics. Per BoE FAQ "Can I store the data
         provided on your website?": "Providing that the data were used
         according to the conditions listed on [the legal page], there would
         be no objection to electronic storage." Citation:
         Bank of England, UK Government Liability Curve — Nominal (daily).
"""

import csv
import io
import re
import sys
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import requests


# Constants
BOE_URL = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/"
    "yield-curves/glcnominalddata.zip"
)
HISTORY_YEARS = 5
TIMEOUT_SECONDS = 120  # ZIP is ~38 MB, allow generous timeout
USER_AGENT = "xccy-g8/1.0 (https://github.com/sanderdayan1982/xccy-g8)"

# Which sheet inside each XLSX holds the spot curve we want
SPOT_CURVE_SHEET = "4. spot curve"

# Tenor in years -> output filename. 0.5 = 6 months.
TENORS = {
    0.5: "GBP_BILL_6M.csv",   # bill_short_GBP for engine
    1.0: "GBP_BILL_1Y.csv",
    2.0: "GBP_BILL_2Y.csv",
    5.0: "GBP_BILL_5Y.csv",
    10.0: "GBP_BILL_10Y.csv",
}

# Match files inside the ZIP. The BoE prefixes filenames with a timestamp;
# we match by the stable "GLC_Nominal_daily_data_*.xlsx" suffix.
XLSX_PATTERN = re.compile(
    r"GLC_Nominal_daily_data_(\d{4})_to_(\d{4}|present)\.xlsx$",
    re.IGNORECASE,
)

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data"


def _select_relevant_xlsx_names(
    zip_namelist: list[str],
    earliest_year_needed: int,
) -> list[str]:
    """
    From all files inside the ZIP, return only those XLSX whose year range
    overlaps with [earliest_year_needed, now]. Saves time and memory by
    skipping the 6+ pre-2016 archives.
    """
    selected: list[str] = []
    current_year = datetime.utcnow().year

    for name in zip_namelist:
        m = XLSX_PATTERN.search(name)
        if not m:
            continue
        start = int(m.group(1))
        end_raw = m.group(2)
        end = current_year if end_raw == "present" else int(end_raw)

        # Include file if it overlaps the needed window
        if end >= earliest_year_needed and start <= current_year:
            selected.append(name)

    return selected


def _find_year_column_indices(
    header_row: tuple,
    wanted_tenors: list[float],
    tolerance: float = 1e-6,
) -> dict[float, int]:
    """
    Inspect the header row (row 4: ['years:', 0.5, 1, 1.5, 2, ...]) and
    return a dict mapping each wanted tenor (in years) to its column index.
    """
    col_for_tenor: dict[float, int] = {}
    for j, cell in enumerate(header_row):
        if cell is None or isinstance(cell, str):
            continue
        try:
            cell_year = float(cell)
        except (TypeError, ValueError):
            continue
        for tenor in wanted_tenors:
            if tenor in col_for_tenor:
                continue
            if abs(cell_year - tenor) < tolerance:
                col_for_tenor[tenor] = j

    return col_for_tenor


def _parse_xlsx_bytes(
    xlsx_bytes: bytes,
    date_from: datetime,
    date_to: datetime,
) -> dict[float, list[tuple[str, float]]]:
    """
    Parse one BoE XLSX (in-memory bytes) and extract rows for the wanted tenors,
    filtered to the [date_from, date_to] window.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), read_only=True, data_only=True
    )

    if SPOT_CURVE_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Expected sheet '{SPOT_CURVE_SHEET}' not found. "
            f"Got: {wb.sheetnames}"
        )

    ws = wb[SPOT_CURVE_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    all_rows = list(rows_iter)
    wb.close()

    if len(all_rows) < 6:
        return {t: [] for t in TENORS}

    # Row 4 (index 3) holds 'years:', 0.5, 1, 1.5, ...
    header_row = all_rows[3]
    if not header_row or header_row[0] != "years:":
        # Fallback: search rows 0..7 for one starting with 'years:'
        header_row = None
        for r in all_rows[:8]:
            if r and r[0] == "years:":
                header_row = r
                break
        if header_row is None:
            raise ValueError(
                f"Could not locate 'years:' header row in '{SPOT_CURVE_SHEET}'"
            )

    col_for_tenor = _find_year_column_indices(header_row, list(TENORS.keys()))

    missing = [t for t in TENORS if t not in col_for_tenor]
    if missing:
        raise ValueError(
            f"Tenors {missing} not found in header row. "
            f"Header sample: {header_row[:15]}"
        )

    results: dict[float, list[tuple[str, float]]] = {t: [] for t in TENORS}

    # Data rows: row 6 onward (index 5+), but be lenient — accept any row
    # whose col[0] is a datetime within the window.
    for raw in all_rows[5:]:
        if not raw or raw[0] is None:
            continue

        # Date is either a datetime object or sometimes a string
        cell0 = raw[0]
        if isinstance(cell0, datetime):
            date_obj = cell0
        else:
            try:
                date_obj = datetime.strptime(str(cell0).strip()[:10], "%Y-%m-%d")
            except ValueError:
                continue

        if date_obj < date_from or date_obj > date_to:
            continue

        date_str = date_obj.strftime("%Y%m%d")

        for tenor, col_idx in col_for_tenor.items():
            if col_idx >= len(raw):
                continue
            cell = raw[col_idx]
            if cell is None or cell == "" or isinstance(cell, str):
                continue
            try:
                value = float(cell)
            except (TypeError, ValueError):
                continue
            results[tenor].append((date_str, value))

    return results


def fetch_gbp_bills(
    date_from: datetime,
    date_to: datetime,
) -> dict[float, list[tuple[str, float]]]:
    """
    Download the BoE nominal yield curve ZIP, extract relevant XLSX files,
    and combine rows from all of them into per-tenor sorted lists.
    """
    headers = {"User-Agent": USER_AGENT, "Accept": "application/zip"}

    print(f"Downloading {BOE_URL}")
    response = requests.get(BOE_URL, headers=headers, timeout=TIMEOUT_SECONDS)
    response.raise_for_status()

    print(f"Downloaded {len(response.content) / 1024 / 1024:.1f} MB. Unzipping...")

    with zipfile.ZipFile(io.BytesIO(response.content)) as zf:
        namelist = zf.namelist()
        relevant = _select_relevant_xlsx_names(namelist, date_from.year)

        if not relevant:
            raise ValueError(
                f"No relevant XLSX found in ZIP. All names: {namelist}"
            )

        print(f"Reading {len(relevant)} XLSX file(s) covering {date_from.year}+:")
        combined: dict[float, list[tuple[str, float]]] = {t: [] for t in TENORS}

        for name in sorted(relevant):
            print(f"  - {name.split('/')[-1]}")
            with zf.open(name) as f:
                xlsx_bytes = f.read()
            partial = _parse_xlsx_bytes(xlsx_bytes, date_from, date_to)
            for tenor, rows in partial.items():
                combined[tenor].extend(rows)

    # Sort by date (each XLSX file is already chronological, but we combined)
    for tenor in combined:
        # Deduplicate in case adjacent files overlap (defensive)
        seen = {}
        for d, v in combined[tenor]:
            seen[d] = v
        combined[tenor] = sorted(seen.items())

    return combined


def write_csv(rows: list[tuple[str, float]], output_path: Path) -> None:
    """Write rows to OHLCV format CSV (O=H=L=C for daily rates, V=0)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["DATE", "OPEN", "HIGH", "LOW", "CLOSE", "VOLUME"])
        for date_str, value in rows:
            v = f"{value:.4f}"
            writer.writerow([date_str, v, v, v, v, "0"])


def main() -> int:
    today = datetime.utcnow()
    date_from = today - timedelta(days=365 * HISTORY_YEARS)

    print(f"Fetching GBP gilt yields from {date_from.date()} to {today.date()}")
    print(f"Source: Bank of England (nominal spot curve)")
    print(f"Tenors: {[f'{t}Y' if t >= 1 else f'{int(t*12)}M' for t in TENORS]}")
    print()

    try:
        results = fetch_gbp_bills(date_from, today)
    except requests.HTTPError as exc:
        print(f"ERROR: BoE HTTP error: {exc}", file=sys.stderr)
        return 1
    except requests.RequestException as exc:
        print(f"ERROR: BoE network error: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"ERROR: GBP bills fetch failed: {exc}", file=sys.stderr)
        return 1

    print()
    successes = 0
    failures = 0

    for tenor, filename in TENORS.items():
        rows = results.get(tenor, [])
        output_path = OUTPUT_DIR / filename

        if not rows:
            label = f"{int(tenor*12)}M" if tenor < 1 else f"{int(tenor)}Y"
            print(f"[{label}] ERROR: No rows returned", file=sys.stderr)
            failures += 1
            continue

        write_csv(rows, output_path)
        label = f"{int(tenor*12)}M" if tenor < 1 else f"{int(tenor)}Y"
        print(f"[{label}] OK: Wrote {len(rows)} rows to {filename}")
        print(f"        Latest:   {rows[-1][0]} = {rows[-1][1]:.4f}%")
        print(f"        Earliest: {rows[0][0]} = {rows[0][1]:.4f}%")
        print()
        successes += 1

    print(f"Summary: {successes} OK, {failures} failed (of {len(TENORS)} total)")
    return 0 if failures == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
